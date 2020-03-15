# -*- coding: utf-8 -*-

from django.core.management.base import BaseCommand, CommandError
from deal.models import Deal, DealPerson, Service
from identity.models import Person, PersonPhone

from openpyxl import load_workbook


class Command(BaseCommand):
    help = 'parse xls'

    def handle(self, *args, **options):
        print('Begin parse\n')

        filename = '/srv/absolutum/other/calls.xlsx'
        wb = load_workbook(filename=filename)
        sheet_pravka = wb['calls']
        phones = []
        for i in range(2, 160):
            phone = sheet_pravka.cell(row=i, column=3).value
            try:
                phone_qs = Person.objects.filter(phones__value=str(phone)).first()
                print(phone, phone_qs.cache.get('full_name'), phone_qs.id)
            except:
                if phone and phone not in phones:
                    print(phone)
                    phones.append(phone)

        print('\n')
        for i in list(set(phones)):
            print(i)

        # Deal.objects.filter(service_type=None).update(service_type=1)

        print('End parse')
