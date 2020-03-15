# -*- coding: utf-8 -*-

from datetime import date, datetime
from django.core.management.base import BaseCommand, CommandError
from deal.models import Deal, DealPerson, Service
from identity.models import Person, PersonPhone

from openpyxl import load_workbook


class Command(BaseCommand):
    help = 'parse xls'

    def handle(self, *args, **options):
        print('Begin parse\n')

        # service = Service.objects.filter(branch__city__name='Краснодар').first()
        service = Service.objects.get(branch__city__name='Нижний Новгород')
        step = service.template.steps.filter(number=1).first()

        filename = '/srv/absolutum/other/dlya_vygruzki_Nizhniy_Novgorod.xlsx'
        wb = load_workbook(filename=filename)

        sheet_control = wb['list1']

        # Контроль
        service_type = service.template.service_types.filter(name='Бесплатно').first()
        for i in range(4, 90):
            person_kwargs = dict()
            fio = sheet_control.cell(row=i, column=2).value
            if not fio:
                break
            fio_split = fio.strip().replace('\xa0', ' ').replace('  ', ' ').split(' ')
            person_kwargs['last_name'] = fio_split[0]
            if len(fio_split) > 1 and fio_split[1]:
                person_kwargs['first_name'] = fio_split[1]
            if len(fio_split) > 2 and fio_split[2]:
                person_kwargs['patronymic'] = fio_split[2]

            person_kwargs['birthday'] = sheet_control.cell(row=i, column=3).value
            # person_kwargs['birthday'] = datetime.strptime(birthday, '%d.%m.%Y').date()
            phone = sheet_control.cell(row=i, column=4).value.replace(' ', '').replace('-', '')
            if phone[0] == '8':
                phone = '7%s' % phone[1:]

            try:
                person = Person.objects.create(**person_kwargs)
                person.phones.create(value=phone)
            except:
                person = None
                print('unique', i, fio_split)

            if person:
                deal = Deal.objects.create(
                    service=service,
                    comment='18-19 мая 2019',
                    service_type=service_type,
                    step=step,
                    cost=service_type.cost
                )
                DealPerson.objects.create(
                    deal=deal,
                    person=person,
                    control=True
                )
                deal.save()
                print(i, fio_split, phone)

        # Deal.objects.filter(service_type=None).update(service_type=1)

        print('End parse')
