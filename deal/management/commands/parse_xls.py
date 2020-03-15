# -*- coding: utf-8 -*-

from django.core.management.base import BaseCommand, CommandError
from deal.models import Deal, DealPerson, Service
from identity.models import Person, PersonPhone

from openpyxl import load_workbook


class Command(BaseCommand):
    help = 'parse xls'

    def handle(self, *args, **options):
        print('Begin parse\n')

        # service = Service.objects.filter(branch__city__name='Краснодар').first()
        service = Service.objects.filter(branch__city__name='Ростов на Дону').first()
        step = service.template.steps.filter(step=1).first()

        # filename = '/srv/absolutum/crm_krasnodar_avgust.xlsx'
        filename = '/srv/absolutum/crm_rostov_avgust.xlsx'
        wb = load_workbook(filename=filename)

        sheet_pravka = wb['pravka']
        sheet_control = wb['control']

        # Правка
        service_type = service.template.service_types.filter(name='Взрослый').first()
        for i in range(4, 90):
            person_kwargs = dict()
            fio = sheet_pravka.cell(row=i, column=2).value
            if not fio:
                break
            fio_split = fio.strip().replace('\xa0', ' ').replace('  ', ' ').split(' ')
            person_kwargs['last_name'] = fio_split[0]
            if len(fio_split) > 1 and fio_split[1]:
                person_kwargs['first_name'] = fio_split[1]
            if len(fio_split) > 2 and fio_split[2]:
                person_kwargs['patronymic'] = fio_split[2]

            age = sheet_pravka.cell(row=i, column=3).value
            phone = sheet_pravka.cell(row=i, column=4).value.replace(' ', '').replace('-', '')
            if phone[0] == '8':
                phone = '7%s' % phone[1:]
            other = sheet_pravka.cell(row=i, column=5).value
            city = sheet_pravka.cell(row=i, column=7).value

            comment_list = []
            if age:
                comment_list.append(('Возвраст: %s' % age).strip())
            if city:
                comment_list.append(('проживают в %s' % city).strip())
            if other:
                comment_list.append(str(other).strip())
            comment = ', '.join(comment_list)

            person = Person.objects.create(**person_kwargs)
            person.phones.create(value=phone)
            deal = Deal.objects.create(
                service=service,
                comment=comment,
                service_type=service_type,
                step=step,
                cost=service_type.cost
            )
            DealPerson.objects.create(
                deal=deal,
                person=person
            )
            print(i, fio_split, phone, comment)

        # Контроль
        service_type = service.template.service_types.filter(name='Бесплатно').first()
        for i in range(4, 90):
            person_kwargs = dict()
            fio = sheet_control.cell(row=i, column=2).value
            if not fio:
                break
            fio_split = fio.strip().replace('\xa0', ' ').replace('  ', ' ').split(' ')
            person_kwargs['last_name'] = fio_split[0]
            if len(fio_split) > 1 and fio_split[1]:
                person_kwargs['first_name'] = fio_split[1]
            if len(fio_split) > 2 and fio_split[2]:
                person_kwargs['patronymic'] = fio_split[2]

            age = sheet_control.cell(row=i, column=3).value
            phone = sheet_control.cell(row=i, column=4).value.replace(' ', '').replace('-', '')
            if phone[0] == '8':
                phone = '7%s' % phone[1:]
            other = sheet_control.cell(row=i, column=5).value
            city = sheet_control.cell(row=i, column=7).value

            comment_list = []
            if age:
                comment_list.append(('Возвраст: %s' % age).strip())
            if city:
                comment_list.append(('проживают в %s' % city).strip())
            if other:
                comment_list.append(str(other).strip())
            comment = ', '.join(comment_list)

            person = Person.objects.create(**person_kwargs)
            person.phones.create(value=phone)
            deal = Deal.objects.create(
                service=service,
                comment=comment,
                service_type=service_type,
                step=step,
                cost=service_type.cost
            )
            DealPerson.objects.create(
                deal=deal,
                person=person,
                control=True
            )
            print(i, fio_split, phone, comment)

        # Deal.objects.filter(service_type=None).update(service_type=1)

        print('End parse')
