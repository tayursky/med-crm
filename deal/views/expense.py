import json
from datetime import datetime, date
from decimal import Decimal

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.writer.excel import save_virtual_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, Side, numbers
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule

from django.core import serializers
from django.db.models import Q
from django.urls import reverse, reverse_lazy
from django.utils.http import urlquote
from django.http import HttpResponse, JsonResponse
from django.views.generic import DetailView, ListView, TemplateView, View
from django.shortcuts import render, get_object_or_404
from django.forms.models import modelform_factory
from django.contrib.auth.mixins import LoginRequiredMixin

from absolutum.mixins import CoreMixin, DisplayMixin
from deal.models import ExpenseType, Expense, DealExpense
from directory.utils import get_model, get_detail_fields_mapping, get_child_list
from directory.forms import FilterForm
from utils.remote_forms.forms import RemoteForm
from utils.choices import get_choices, filters_choices


class ExpenseDetailView(LoginRequiredMixin, DetailView):
    template_name = 'model_detail.jinja2'

    def dispatch(self, request, *args, **kwargs):
        self.model = Expense
        return super().dispatch(request, *args, **kwargs)

    def get_context_data(self, *, object_list=None, **kwargs):
        context = super().get_context_data()
        context.update({
            'url': self.request.path,
            'title': self.model._meta.verbose_name,
            'model_name': self.model.__name__.lower(),
            'detail_list': get_detail_fields_mapping(context['object']),
            'child_list': get_child_list(self.object),
            'bread_crumbs': self.object.get_bread_crumbs()
        })
        return context


class ExpenseListView(LoginRequiredMixin, ListView, CoreMixin):
    model = DealExpense
    action = None
    object = None
    related_name = None
    list_display = None
    count = 0
    total = 0
    permissions = []
    filters = {}
    filters_q = {}

    def dispatch(self, request, *args, **kwargs):
        self.permissions = DealExpense.get_permissions(request)
        self.model = Expense if kwargs.get('model_name') == 'expense' else DealExpense
        self.filters = {}
        self.filters = self.model.get_filters(request)
        self.filters = filters_choices(request, self.filters, self.model)
        self.filters_q = self.model.get_filters_q(self.request, filters=self.filters)
        self.action = kwargs.get('action')
        self.permissions = self.model.get_permissions(request)
        self.list_display = ['id'] + getattr(self.model, 'list_display', [])
        return super().dispatch(request, *args, **kwargs)

    def get_queryset(self):
        queryset = self.model.objects.all()
        list_related = getattr(self.model, 'list_related', [])
        if self.object and self.related_name:
            queryset = getattr(self.object, self.related_name).all()
        queryset = queryset.filter(self.filters_q).select_related(*list_related)
        return queryset

    def get_items(self):
        items = []
        request_dict = self.request.GET.dict()
        request_dict.pop('get', None)
        if len(request_dict.keys()) <= 0:
            return items

        expense_q = self.get_queryset()

        # Фильтр только по ответственным филиалам
        if not self.request.user.has_perm('Администраторы'):
            expense_q = expense_q.filter(
                branch__in=get_choices(self.request, 'company.Branch', get_list=True),
                created_by=self.request.user.person
            ).exclude(
                type__name='Заработная плата'
            )

        self.count = expense_q.count()
        for q_item in expense_q:
            item = dict()
            for field_name in self.list_display + ['description']:
                _q_item = q_item
                for field in field_name.split('__'):
                    try:
                        _q_item = getattr(q_item, 'get_%s_display' % field)()
                    except AttributeError:
                        _q_item = getattr(_q_item, field)
                item[field_name] = _q_item.__str__() if _q_item else ''
            self.total += q_item.value
            items.append(item)
        return items

    def get(self, request, *args, **kwargs):
        if 'view' not in self.permissions:
            return JsonResponse(dict(answer='No permissions'), safe=False)
        if not request.GET:
            return render(request, 'app_vue.jinja2')
        elif self.action == 'get_xls':
            return self.get_xls()
        context = dict(
            title=self.model._meta.verbose_name_plural,
            url=self.request.path,
            child_list=get_child_list(self.object),
            pk=self.object.pk if self.object else '',
            related_name=self.related_name,
            actions=self.model.base_actions,
            headers=self.model.get_headers(),
            items=self.get_items(),
            count=self.count,
            total=str(self.total),
            permissions=self.permissions,
            filters=self.filters
        )
        if request.GET.get('debug'):
            from django.shortcuts import render_to_response
            return render_to_response('debug.jinja2', locals())

        return JsonResponse(context)

    def get_xls(self):
        filename = 'Расходы.xlsx'.format(
            service='self.service',
        )
        title_text = 'Расходы'
        wb = Workbook()
        ws = wb.active
        ws.title = '%s %s' % ('Расходы', '')

        title = ws.cell(row=1, column=1, value=title_text)
        title.alignment = Alignment(horizontal='left')
        title.font = Font(name='Calibri', size=11, bold=True)
        ws.merge_cells('A1:G1')

        labels = ['Филиал', 'День', 'Тип расходов', 'Другой тип расходов', 'Описание', 'Создал запись', 'Сумма']
        columns = []
        row = 3
        for index, label in enumerate(labels):
            cell = ws.cell(row=row, column=index + 1, value=label)
            cell.font = Font(name='Calibri', size=11, bold=True)
            columns.append(cell.column)

        summary = Decimal(0.0)
        for item in self.get_items():
            # expense = Expense.objects.get(pk=item['id'])
            row += 1
            summary += Decimal(item['value'])
            _ = ws.cell(row=row, column=1, value=item['branch'])
            _ = ws.cell(row=row, column=2, value=item['day'])
            _ = ws.cell(row=row, column=3, value=item['type'])
            _ = ws.cell(row=row, column=4, value=item['other_type'])
            _ = ws.cell(row=row, column=5, value=item['description'])
            _ = ws.cell(row=row, column=6, value=item['created_by'])
            _ = ws.cell(row=row, column=7, value=Decimal(item['value']))
            _.number_format = numbers.FORMAT_NUMBER_00

        row += 2
        _ = ws.cell(row=row, column=6, value='Итого:')
        _ = ws.cell(row=row, column=7, value=str(summary))

        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 20
        ws.column_dimensions['F'].width = 30
        ws.column_dimensions['G'].width = 15

        response = HttpResponse(save_virtual_workbook(wb), content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'inline; filename=%s' % urlquote(filename).lower()
        return response
