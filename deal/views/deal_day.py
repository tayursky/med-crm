import calendar
import decimal
import math
from datetime import date, datetime, timedelta

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.writer.excel import save_virtual_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, Side
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule

from django.http import HttpResponse, JsonResponse
from django.views.generic import View, ListView
from django.utils.http import urlquote
from django.utils.translation import ugettext as _
from django.shortcuts import render, redirect, render_to_response

from company.models import User, TimeTable
from identity.models import Account
from identity.utils import LoginRequiredMixin
from deal.models import Deal, Service
from utils.date_time import delta_minutes, get_week_start
from deal.views.deal_calendar import DealCalendar


class DayView(LoginRequiredMixin, DealCalendar, View):
    """
        Расписание на день
    """
    day_start = None
    day_finish = None
    timezone = 0
    action = None
    permissions = []

    def dispatch(self, request, *args, **kwargs):
        self.permissions = Deal.get_permissions(request)
        self.service, self.service_set = self.set_service()
        self.timezone = self.service_set['timezone']
        self.action = kwargs.get('action')
        try:
            self.day_start = datetime.strptime(request.GET.get('day'), '%Y%m%d')
        except TypeError:
            pass

        get_day = request.GET.get('get_day')
        if not self.day_start or get_day == 'current':
            self.day_start = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)

        if get_day == 'prev':
            self.day_start -= timedelta(days=1)
        elif get_day == 'next':
            self.day_start += timedelta(days=1)

        # self.day_start += timedelta(hours=self.timezone)
        self.day_finish = self.day_start + timedelta(hours=24)
        self.deals = self.get_deals()
        return super().dispatch(request, *args, **kwargs)

    def get_deals(self):
        deal_queryset = Deal.objects.filter(
            service=self.service,
            start_datetime__gte=self.day_start,
            start_datetime__lte=self.day_finish + timedelta(hours=1),
            step__number__gt=0
        )
        deals = dict()
        values = ['id', 'service__id', 'step__id', 'step__step', 'cost', 'start_datetime', 'finish_datetime']
        for deal in deal_queryset:  # .values(*values):
            persons = deal.get_persons()
            deal.start_datetime += timedelta(hours=self.timezone)
            deal.finish_datetime += timedelta(hours=self.timezone)
            start = int(deal.start_datetime.strftime('%Y%m%d%H%M'))
            finish = int(deal.finish_datetime.strftime('%Y%m%d%H%M'))
            manager = User.objects.filter(account=deal.history.filter(history_type='+').first().history_user_id).first()
            deals[start] = dict(
                persons=persons,
                pravka=min([i['pravka'] for i in persons or [{'pravka': 0}]]),
                id=deal.id,
                service=deal.service_id,
                step=deal.step_id,
                step_number=deal.step.number,
                step_label=deal.step.name,
                cost=deal.cost,
                paid=deal.paid,
                paid_non_cash=deal.paid_non_cash,
                arrear=int(deal.cost - (deal.paid + deal.paid_non_cash)),  # Остаток
                start=start,
                start_string=deal.start_datetime.strftime('%H:%M'),
                finish=finish,
                finish_string=deal.finish_datetime.strftime('%H:%M'),
                manager__full_name='%s %s' % (manager.last_name, manager.first_name) if manager else '',
                master__full_name='%s %s' % (deal.master.last_name, deal.master.first_name) if deal.master else '',
                minutes=delta_minutes(start, finish),
                comment=deal.comment,
            )
        return deals

    def get_day_set(self):
        title = '{day} {month} {year}'.format(
            day=self.day_start.strftime('%d'),
            month=_(self.day_start.strftime('%B')).lower(),
            year=self.day_start.strftime('%Y'),
        )
        timing, groups = self.get_day_timing(self.deals, self.day_start.date())
        timetable = TimeTable.objects.filter(branch=self.service.branch,
                                             plan_start_datetime__gt=self.day_start,
                                             plan_end_datetime__lt=self.day_finish + timedelta(hours=1),
                                             user__account__groups__name='Правщики') \
            .order_by('user__service_default_master', 'user__last_name')
        try:
            master = timetable.first().user
        except AttributeError:
            master = None
        return dict(
            title=title,
            year=self.day_start.year,
            month=self.day_start.month,
            label=self.day_start.strftime('%d.%m.%Y'),
            key=self.day_start.strftime('%Y%m%d'),
            groups=groups,
            timing=timing,
            timing_sorted=sorted(timing.keys()),
            masters=master.last_name if master else '',
            master_full_name=master.get_full_name_display() if master else ''
        )

    def get_perm(self, *args, **kwargs):
        from django.contrib.auth.models import Group, Permission
        result = []
        values = 'id', 'name', 'codename', 'content_type__app_label', 'content_type__model'
        for perm in Permission.objects.all().values(*values):
            result.append(_(perm['name']))
        return result

    def get(self, request, *args, **kwargs):
        if 'view' not in self.permissions:
            return JsonResponse(dict(answer='No permissions'), safe=False)

        if self.action == 'get_xls':
            return self.get_xls()

        if not request.GET:
            return render(request, 'app_vue.jinja2')

        context = dict(
            service_set=self.service_set,
            deals=self.deals,
            day_set=self.get_day_set(),
            # timing=self.get_range(self.day_start),
            # perms=self.get_perm()
        )
        if self.request.GET.get('debug'):
            return render_to_response('debug.jinja2', locals())
        return JsonResponse(context, safe=False)

    def get_xls(self):
        day_set = self.get_day_set()
        deals = self.get_deals()
        filename = '{branch} {service} {day}.xlsx'.format(
            # city=self.service.branch.city.name,
            branch=self.service.branch.name,
            service=self.service.name,
            day=self.day_start.strftime('%Y%m%d')
        )
        wb = Workbook()
        ws = wb.active
        ws.title = '%s %s' % (self.service.name, self.day_start.strftime('%Y%m%d'))

        title_text = '{day} {branch} {service}'.format(
            # city=self.service.branch.city.name,
            branch=self.service.branch.name,
            service=self.service.name,
            day=self.day_start.strftime('%d.%m.%Y')
        )
        title = ws.cell(row=1, column=1, value=title_text)
        title.alignment = Alignment(horizontal='left')
        title.font = Font(name='Calibri', size=11, bold=True)
        ws.merge_cells('A1:C1')

        labels = ['Время', 'Клиент', 'День рождения', 'Телефон', 'Email',
                  'Статус', 'Стоимость', 'Предоплата', 'Безнал',
                  'Правщик', 'Администратор', 'Комментарий']
        # names = [i[1] for i in data['tabs'][tab_key]['labels']]
        columns = []
        for index, label in enumerate(labels):
            cell = ws.cell(row=2, column=index + 1, value=label)
            cell.font = Font(name='Calibri', size=11, bold=True)
            columns.append(cell.column)

        row = 3
        # import ipdb; ipdb.set_trace()
        for time_key in sorted(day_set['timing'].keys()):
            item = day_set['timing'][time_key]
            # time_key = time['key']
            time_key_str = '%s%s' % (day_set['key'], time_key)

            if time_key in day_set['groups']:
                row += 1
                _ = ws.cell(row=row, column=1, value=day_set['groups'][time_key]['name'])
                _ = ws.cell(row=row, column=2, value=day_set['groups'][time_key]['persons'])

            # has deals
            if time_key in day_set['timing'] and 'deals' in item:
                # print('item', item)
                if item.get('empty'):
                    row += 1
                    _ = ws.cell(row=row, column=1, value=item['label'])

                for deal_key in item['deals']:
                    try:
                        deal = deals[deal_key]
                    except KeyError:
                        continue

                    # row += 1
                    # deal_text = '{step} ({cost} / {paid}), правщик: {master}'.format(
                    #     step=deal['step_label'],
                    #     cost=deal['cost'],
                    #     paid=deal['paid'],
                    #     master=deal['master__full_name'],
                    # )
                    # _ = ws.cell(row=row, column=1, value='%s (%s)' % (deal['start_string'], deal['minutes']))
                    # _ = ws.cell(row=row, column=2, value=deal_text)
                    # ws.merge_cells('B{row}:C{row}'.format(row=row))
                    # if deal['comment']:
                    #     row += 1
                    #     _ = ws.cell(row=row, column=2, value=deal['comment'])
                    #     ws.merge_cells('B{row}:C{row}'.format(row=row))

                    for index, person in enumerate(deal['persons']):
                        row += 1
                        person_text = '%s' % (person['full_name'])
                        if person['control']:
                            person_text = '(Контроль) ' + person_text
                        if index == 0:
                            _ = ws.cell(row=row, column=1, value='%s (%s)' % (deal['start_string'], deal['minutes']))
                            _ = ws.cell(row=row, column=6, value=deal['step_label'])
                            _ = ws.cell(row=row, column=7, value=deal['cost'])
                            _ = ws.cell(row=row, column=8, value=deal['paid'])
                            _ = ws.cell(row=row, column=9, value=deal['paid_non_cash'])
                            _ = ws.cell(row=row, column=10, value=deal['master__full_name'])
                            _ = ws.cell(row=row, column=11, value=deal['manager__full_name'])
                            _ = ws.cell(row=row, column=12, value=deal['comment'])
                        _ = ws.cell(row=row, column=2, value=person_text)
                        _ = ws.cell(row=row, column=3, value=person['age'])
                        _ = ws.cell(row=row, column=4, value=person['phone'])
                        # _ = ws.cell(row=row, column=5, value=person['emails'])

                if 'empty_finish' in item:
                    row += 1
                    _ = ws.cell(row=row, column=1, value=item['empty_finish']['label'])

        border = Border(bottom=Side(border_style='thin', color='000000'))
        ws.conditional_formatting.add('A1:K%s' % row, FormulaRule(formula=['E1=0'], border=border))
        ws.column_dimensions['A'].width = 13
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 10

        # row += 2
        # for cellObj in ws['%s2:%s%s' % (columns[0], columns[-1], items_count)]:
        #     for cell in cellObj:
        # import ipdb; ipdb.set_trace()
        # print(cell.coordinate, cell.column)

        # ws.column_dimensions[cell.column].bestFit = True
        # ws[cell.coordinate].alignment = Alignment(horizontal='left')

        response = HttpResponse(save_virtual_workbook(wb), content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'inline; filename=%s' % urlquote(filename).lower()
        return response
