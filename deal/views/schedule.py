import calendar
import decimal
import math
import json
from datetime import date, datetime, timedelta

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.writer.excel import save_virtual_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, Side
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule

from django.http import HttpResponse, JsonResponse
from django.db.models import Case, Subquery, Q, Sum, Avg, Max, F, OuterRef, Value, DecimalField, When
from django.views.generic import View, ListView
from django.utils.http import urlquote
from django.utils.translation import ugettext as _
from django.shortcuts import render, redirect, render_to_response

from company.models import Branch, TimeTable, User
from identity.models import Person
from identity.utils import LoginRequiredMixin
from deal.models import Deal, DealPerson, Service
from identity.models import Person
from utils.date_time import delta_minutes, get_week_start
from deal.views.deal_calendar import DealCalendar


class ScheduleView(LoginRequiredMixin, DealCalendar, View):
    """
        Расписание на неделю
    """
    days, timezone, start, end, today = 0, 0, None, None, None
    permissions = []
    masters = []

    def dispatch(self, request, *args, **kwargs):
        self.today = date.today()
        self.permissions = Deal.get_permissions(request)
        self.days = int(request.GET.get('days', 7))
        self.branch = Branch.objects.get(pk=request.GET.get('branch', 65))
        self.timezone = self.branch.city.timezone
        try:
            self.masters = User.objects.filter(pk__in=json.loads(request.GET.get('masters')))
        except (TypeError, json.decoder.JSONDecodeError):
            pass
        try:
            self.start = datetime.strptime(request.GET.get('current_day'), '%Y-%m-%d').date()
        except (ValueError, TypeError):
            self.start = self.today
        steps = dict(prev=-self.days, next=self.days)
        step = request.GET.get('step')
        if step == 'current':
            self.start = self.today
        else:
            self.start += timedelta(days=steps.get(request.GET.get('step'), 0))

        # Если неделя (7 дней), то первый день берем понедельник
        if self.days == 7:
            self.start = get_week_start(self.start)

        self.start += timedelta(hours=self.timezone)
        self.end = self.start + timedelta(days=self.days)

        print('+++ 0')
        for deal in Deal.objects.filter(services=2):
            print('+++', deal)

        return super().dispatch(request, *args, **kwargs)

    def get_time_groups(self):
        groups = []
        for group in self.branch.time_groups.filter(
                Q(start_date__gte=self.start, start_date__lte=self.end) |
                Q(end_date__gte=self.start, end_date__lte=self.end) |
                Q(start_date__lte=self.start, end_date__gte=self.end)
        ):
            groups.append(dict(
                id=group.id,
                branch=group.branch_id,
                name=group.name,
                users=[u.id for u in group.users.all()],
                timeout=group.timeout,
                start_date=group.start_date,
                end_date=group.end_date,
                start_time=group.start_time.strftime('%H:%M'),
                end_time=group.end_time.strftime('%H:%M'),
            ))
        return groups

    def get_deals(self):
        deal_queryset = Deal.objects.filter(
            branch=self.branch,
            start_datetime__gte=self.start,
            start_datetime__lte=self.end + timedelta(hours=self.timezone),
            stage__step__gt=0
        ).select_related('stage')

        deals = dict()
        for deal in deal_queryset:
            start_iso = (deal.start_datetime + timedelta(hours=self.timezone)).strftime('%Y-%m-%dT%H:%M')
            end_iso = (deal.finish_datetime + timedelta(hours=self.timezone)).strftime('%Y-%m-%dT%H:%M')
            deals[deal.id] = dict(
                id=deal.id,
                branch=deal.branch_id,
                master=deal.master_id,
                manager=deal.manager_id,
                stage=deal.stage.step,
                stage_step=deal.stage.step,
                cost=deal.cost,
                paid=deal.paid,
                paid_non_cash=deal.paid_non_cash,
                start_iso=start_iso,
                end_iso=end_iso,
                minutes=delta_minutes(start_iso, end_iso),
                comment=deal.comment,
                pravka=deal.cache.get('pravka'),
                persons=[]
            )
        for rel in DealPerson.objects.filter(deal__in=[i for i in deals.keys()]).select_related('person'):
            deals[rel.deal_id]['persons'].append(dict(
                control=rel.control,
                primary=rel.primary,
                id=rel.person.id,
                cache=rel.person.cache
            ))

        return [v for k, v in deals.items()]

    def get_schedule(self):
        _days = []
        timetable_qs = self.branch.timetables.filter(
            plan_start_datetime__gt=self.start,
            plan_end_datetime__lt=self.end,
            user__account__groups__name='Правщики'
        ).order_by(
            'plan_start_datetime',
            'user__last_name'
        ).select_related('user')
        timetable = dict()
        for item in timetable_qs:
            _day = item.plan_start_datetime.date()
            if _day not in timetable:
                timetable[_day] = dict(masters=[])
            user = getattr(item, 'user')
            timetable[_day]['masters'].append(dict(
                id=user.id,
                last_name=getattr(user, 'last_name'),
                full_name=user.get_full_name_display(),
                plan_start_iso=(item.plan_start_datetime + timedelta(hours=self.timezone)).strftime('%Y-%m-%d %H:%M'),
                plan_end_iso=(item.plan_end_datetime + timedelta(hours=self.timezone)).strftime('%Y-%m-%d %H:%M'),
            ))
        for day, weekday in enumerate(range(self.days)):
            _day = self.start + timedelta(days=day)
            _days.append(dict(
                iso=_day.strftime('%Y-%m-%d'),
                month_name=_(_day.strftime('%B')),
                weekday=weekday,
                weekday_label=_(_day.strftime('%a')),
                masters=timetable.get(_day, {}).get('masters', []),
            ))

        title_days = str(int(self.start.strftime('%d')))
        end_title = self.end - timedelta(days=1)
        if self.start != end_title:
            if self.start.strftime('%B') != end_title.strftime('%B'):
                title_days += ' %s' % _(self.start.strftime('%B')).lower()
            if self.start.strftime('%Y') != end_title.strftime('%Y'):
                title_days += ' %s' % _(self.start.strftime('%Y'))
            title_days += ' – %s' % int(end_title.strftime('%d'))
        title = '{title_days} {month} {year}'.format(
            title_days=title_days,
            month=_(end_title.strftime('%B')).lower(),
            year=end_title.strftime('%Y'),
        )
        return dict(
            today=self.today.strftime('%Y-%m-%d'),
            title=title,
            start_iso=self.start.strftime('%Y-%m-%d'),
            days=_days,
            range=self.get_range(self.start, self.end),
            time_groups=self.get_time_groups(),
        )

    def get(self, request, *args, **kwargs):
        if 'view' not in self.permissions:
            return JsonResponse(dict(answer='No permissions'), safe=False)
        if not request.GET:
            return render(request, 'app_vue.jinja2')
        context = dict(
            current_day=self.start.strftime('%Y-%m-%d'),
            deals=self.get_deals(),
            branch=self.branch.id,
            schedule=self.get_schedule(),
            stages=self.stages
        )
        if self.request.GET.get('debug'):
            return render_to_response('debug.jinja2', locals())
        return JsonResponse(context, safe=False)


class ScheduleXlsView(ScheduleView):

    def get(self, request, *args, **kwargs):
        deals = self.get_deals()
        filename = '{day}_{branch}_{master}.xlsx'.format(
            day=self.start.strftime('%Y-%m-%d'),
            branch=self.branch.name,
            master=self.masters[0].last_name
        )
        wb = Workbook()
        ws = wb.active
        ws.title = '%s %s' % (self.branch.name, self.start.strftime('%Y%m%d'))

        title_text = '{day} {branch}'.format(
            branch=self.branch.name,
            day=self.start.strftime('%d.%m.%Y')
        )
        title = ws.cell(row=1, column=1, value=title_text)
        title.alignment = Alignment(horizontal='left')
        title.font = Font(name='Calibri', size=11, bold=True)
        ws.merge_cells('A1:C1')

        labels = ['Время', 'Клиент', 'День рождения', 'Телефон', 'Email', 'Статус', 'Стоимость', 'Предоплата', 'Безнал',
                  'Комментарий']
        # names = [i[1] for i in data['tabs'][tab_key]['labels']]
        columns = []
        for index, label in enumerate(labels):
            cell = ws.cell(row=2, column=index + 1, value=label)
            cell.font = Font(name='Calibri', size=11, bold=True)
            columns.append(cell.column)

        row = 3
        for deal in sorted(self.get_deals(), key=lambda x: x['start_iso']):

            # has deals
            for index, person in enumerate(deal['persons']):
                row += 1
                person_text = '%s' % (person['cache'].get('full_name'))
                if person['control']:
                    person_text = '(Контроль) ' + person_text
                if index == 0:
                    _ = ws.cell(row=row, column=1, value='%s (%s)' % (deal['start_iso'][11:17], deal['minutes']))
                    _ = ws.cell(row=row, column=6, value=self.stages[deal['stage']]['label'])
                    _ = ws.cell(row=row, column=7, value=deal['cost'])
                    _ = ws.cell(row=row, column=8, value=deal['paid'])
                    _ = ws.cell(row=row, column=9, value=deal['paid_non_cash'])
                    _ = ws.cell(row=row, column=10, value=deal['comment'])
                _ = ws.cell(row=row, column=2, value=person_text)
                _ = ws.cell(row=row, column=3, value=person['cache'].get('age', ''))
                _ = ws.cell(row=row, column=4, value=person['cache'].get('phone', ''))
                _ = ws.cell(row=row, column=5, value=person['cache'].get('emails', ''))

        border = Border(bottom=Side(border_style='thin', color='000000'))
        ws.conditional_formatting.add('A1:K%s' % row, FormulaRule(formula=['E1=0'], border=border))
        ws.column_dimensions['A'].width = 13
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 10

        # row += 2
        # for cellObj in ws['%s2:%s%s' % (columns[0], columns[-1], items_count)]:
        #     for cell in cellObj:
        # import ipdb; ipdb.set_trace()
        # print(cell.coordinate, cell.column)

        # ws.column_dimensions[cell.column].bestFit = True
        # ws[cell.coordinate].alignment = Alignment(horizontal='left')

        response = HttpResponse(save_virtual_workbook(wb), content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'inline; filename=%s' % urlquote(filename).lower()
        return response
