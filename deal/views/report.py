import json
from datetime import datetime, date, timedelta
from decimal import Decimal

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.writer.excel import save_virtual_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, Side, numbers
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule

from django.core import serializers
from django.contrib.auth.mixins import LoginRequiredMixin
from django.db.models import Q
from django.urls import reverse, reverse_lazy
from django.utils.http import urlquote
from django.http import HttpResponse, JsonResponse
from django.views.generic import DetailView, ListView, TemplateView, View
from django.shortcuts import render, get_object_or_404
from django.forms.models import modelform_factory
from django.db.models import Sum

from absolutum.mixins import CoreMixin, DisplayMixin
from directory.forms import FilterForm
from directory.utils import get_model, get_detail_fields_mapping, get_child_list
from deal.models import Deal, Expense, Report, Service
from utils.remote_forms.forms import RemoteForm
from utils.choices import get_choices


class ReportListView(LoginRequiredMixin, ListView, CoreMixin):
    model = Report
    action = None
    object = None
    related_name = None
    list_display = None
    count = 0
    perms = []
    permission_q = Q()
    summary = dict()
    filters = dict()
    permissions = []

    def dispatch(self, request, *args, **kwargs):
        self.permissions = Deal.get_permissions(request)
        self.action = kwargs.get('action')
        self.list_display = ['id'] + getattr(self.model, 'list_display', [])
        self.perms = self.model.get_permissions(request)
        self.summary = dict(
            ordered=['deals', 'done', 'cancel', 'new', 'control', 'cost', 'paid', 'paid_non_cash'],
            labels=dict(deals='Сделок', done='Успешных', cancel='Отмененных', new='Правок', control='Контролей',
                        cost='На сумму', paid='Всего оплачено', paid_non_cash='В том числе безнал'),
            values=dict(deals=0, done=0, cancel=0, new=0, control=0,
                        cost=Decimal(0.0), paid=Decimal(0.0), paid_non_cash=Decimal(0.0))
        )
        self.filters = self.set_filters(request)

        if not request.user.has_perm('Администраторы'):
            person_id = request.user.person.id
            self.permission_q = Q(branch__managers=person_id)

        return super().dispatch(request, *args, **kwargs)

    def set_filters(self, request):
        _start = date.today().replace(day=1)
        _end = (_start + timedelta(days=33)).replace(day=1) - timedelta(days=1)
        filters = dict(
            data=dict(
                start_datetime=[_start.strftime('%d.%m.%Y'), _end.strftime('%d.%m.%Y')],
                status='closed'
            ),
            ordered=['branch', 'start_datetime', 'manager', 'master', 'persons__control', 'status'],
            fields=dict(
                branch=dict(
                    label='Филиал', key='branch__in',
                    widget=dict(
                        attrs={}, name='SelectMultiple', input_type='select',
                        choices=get_choices(request, 'company.Branch')
                    )
                ),
                start_datetime=dict(
                    label='Период', key='start_datetime',
                    widget=dict(
                        attrs={}, name="DateInput", input_type="daterange"
                    )
                ),
                manager=dict(
                    label='Организатор', key='manager',
                    widget=dict(
                        attrs={}, name="Select", input_type="select", choices=get_choices(request, 'company.Manager')
                    )
                ),
                master=dict(
                    label='Правщик', key='master',
                    widget=dict(
                        attrs={}, name="Select", input_type="select", choices=get_choices(request, 'company.Master')
                    )
                ),
                persons__control=dict(
                    label='Тип услуги', key='rel_persons__control',
                    widget=dict(
                        attrs={}, name="Select", input_type="select",
                        choices=[dict(label="Правка", value=False), dict(label="Контроль", value=True)]
                    )
                ),
                status=dict(
                    label='Статус сделки', key='status',
                    widget=dict(
                        attrs={}, name="Select", input_type="select",
                        choices=[dict(label="Завершенные", value='closed'), dict(label="В работе", value='in_work')]
                    )
                ),
            )
        )
        return self.get_filters(request=request, filters=filters)

    def get_queryset(self):
        list_related = getattr(self.model, 'list_related', [])
        queryset = self.model.objects \
            .filter(self.model.get_filters_q(self.request, filters=self.filters), self.permission_q) \
            .select_related(*list_related)
        return queryset.distinct()

    def get_items(self):
        items = []

        request_dict = self.request.GET.dict()
        request_dict.pop('get', None)
        if not request_dict.get('start_datetime') or len(request_dict.keys()) <= 0:
            return items

        report_q = self.get_queryset()

        self.count = report_q.count()
        for q_item in report_q:
            item = {}
            for field_name in self.list_display:
                _q_item = q_item
                for field in field_name.split('__'):
                    if hasattr(q_item, 'get_%s_display' % field):
                        _q_item = getattr(q_item, 'get_%s_display' % field)()
                    try:
                        _q_item = getattr(_q_item, field)
                    except AttributeError:
                        pass
                item[field_name] = _q_item.__str__() if _q_item else ''

            self.summary['values']['deals'] += 1
            self.summary['values']['cost'] += q_item.cost
            if item['paid'] or item['paid_non_cash']:
                paid = Decimal(item.get('paid') or 0.0) + Decimal(item.get('paid_non_cash') or 0.0)
                self.summary['values']['paid'] += paid
                self.summary['values']['paid_non_cash'] += Decimal(item.get('paid_non_cash') or 0.0)

            if q_item.stage.name == 'cancel':
                self.summary['values']['cancel'] += 1
            elif q_item.stage.name == 'done':
                self.summary['values']['done'] += 1
                for person in q_item.rel_persons.all():
                    if person.control:
                        self.summary['values']['control'] += 1
                    else:
                        self.summary['values']['new'] += 1

            items.append(item)

        return items

    def get(self, request, *args, **kwargs):
        if 'view' not in self.permissions:
            return JsonResponse(dict(answer='No permissions'), safe=False)
        if not request.GET:
            return render(request, 'app_vue.jinja2')
        elif self.action == 'get_xls':
            return self.get_xls()

        context = dict(
            title=self.model._meta.verbose_name_plural,
            url=self.request.path,
            child_list=get_child_list(self.object),
            pk=self.object.pk if self.object else '',
            related_name=self.related_name,
            headers=self.model.get_headers(),
            filters=self.filters,
            count=self.count,
            items=self.get_items(),
            summary=self.summary,
            permissions=self.perms,
        )

        if request.GET.get('debug'):
            from django.shortcuts import render_to_response
            return render_to_response('debug.jinja2', locals())

        return JsonResponse(context)

    def get_xls(self):
        filename = 'Отчет.xlsx'.format(
            service='self.service',
        )
        title_text = 'Отчет'.format(
            service='self.service.name',
        )

        expense_q = Expense.objects.all()
        expense_q = expense_q.filter(branch__in=self.filters['data'].get('branch', []))
        if self.request.GET.get('start_datetime[]'):
            json_str = str(self.request.GET).replace('<QueryDict: ', '').replace('>', '')
            daterange = eval(json_str)['start_datetime[]']
            daterange_start = datetime.strptime(daterange[0], '%d.%m.%Y')
            daterange_finish = datetime.strptime(daterange[1], '%d.%m.%Y').replace(hour=23, minute=59)
            expense_q = expense_q.filter(**dict(
                day__gte=daterange_start, day__lte=daterange_finish
            ))
        expense_sum = expense_q.aggregate(Sum('value'))['value__sum']
        expense_sum = expense_sum if expense_sum else Decimal('0.00')

        wb = Workbook()
        ws = wb.active
        ws.title = '%s %s' % ('Отчет', '')

        title = ws.cell(row=1, column=1, value=title_text)
        title.alignment = Alignment(horizontal='left')
        title.font = Font(name='Calibri', size=11, bold=True)
        ws.merge_cells('A1:G1')

        labels = ['id', 'Город', 'День', 'Организатор', 'Правщик', 'Клиент', 'Правка',
                  'Стоимость', 'Оплачено', 'Безнал', 'Статус']
        columns = []
        row = 7
        for index, label in enumerate(labels):
            cell = ws.cell(row=row, column=index + 1, value=label)
            cell.font = Font(name='Calibri', size=11, bold=True)
            columns.append(cell.column)

        for item in self.get_items():
            try:
                deal = Deal.objects.get(pk=item['id'])
            except Deal.DoesNotExist:
                continue
            persons = deal.get_persons()
            row += 1
            _ = ws.cell(row=row, column=1, value=item['id'])
            _ = ws.cell(row=row, column=2, value=item['service'])
            _ = ws.cell(row=row, column=3, value=item['day'])
            _ = ws.cell(row=row, column=4, value=item['manager'])
            _ = ws.cell(row=row, column=5, value=item['master'])

            _ = ws.cell(row=row, column=8, value=float(item.get('cost') or 0))
            _.number_format = numbers.FORMAT_NUMBER_00
            _ = ws.cell(row=row, column=9, value=float(item.get('paid') or 0))
            _.number_format = numbers.FORMAT_NUMBER_00
            _ = ws.cell(row=row, column=10, value=float(item.get('paid_non_cash') or 0))
            _.number_format = numbers.FORMAT_NUMBER_00

            status = deal.stage.name if deal.step.name in ['cancel', 'done'] else 'В работе'
            _ = ws.cell(row=row, column=11, value=status)

            for person in persons:
                _ = ws.cell(row=row, column=6, value=person['full_name'])
                _ = ws.cell(row=row, column=7, value='Правка')
                if person['control']:
                    _ = ws.cell(row=row, column=7, value='Контроль')
                row += 1
            row -= 1
        summary = 'Оплачено: {paid}, Сделок: {deals}, Успешных: {done}, Отмененных: {cancel}, ' \
                  'Правок: {new}, Контролей: {control}' \
            .format(
            paid=self.summary['values']['paid'],
            deals=self.summary['values']['deals'],
            done=self.summary['values']['done'],
            cancel=self.summary['values']['cancel'],
            control=self.summary['values']['control'],
            new=self.summary['values']['new']
        )
        _ = ws.cell(row=3, column=1, value=summary)
        ws.merge_cells('A1:J1')
        ws.merge_cells('A2:J2')
        ws.merge_cells('A3:J3')
        ws.merge_cells('A4:J4')
        ws.merge_cells('A5:J5')
        ws.merge_cells('A6:J6')
        # border = Border(bottom=Side(border_style='thin', color='000000'))
        # ws.conditional_formatting.add('A1:G%s' % row, FormulaRule(formula=['E1=0'], border=border))
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 20
        ws.column_dimensions['F'].width = 30
        ws.column_dimensions['G'].width = 10
        ws.column_dimensions['H'].width = 15
        ws.column_dimensions['I'].width = 15
        ws.column_dimensions['J'].width = 15
        ws.column_dimensions['K'].width = 20

        _ = ws.cell(row=4, column=1, value='Расходы: %s' % to_int(expense_sum))
        _ = ws.cell(row=5, column=1,
                    value='Итого: %s' % to_int(self.summary['values']['paid'] - expense_sum))

        response = HttpResponse(save_virtual_workbook(wb), content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'inline; filename=%s' % urlquote(filename).lower()
        return response


def to_int(value):
    return str(value).replace('.', ',').replace("'", '')
